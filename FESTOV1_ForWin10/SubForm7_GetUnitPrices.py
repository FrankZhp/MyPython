#! -*- coding utf-8 -*-
#! @Time  :2019/3/24 9:00
#! Author :Frank Zhang
#! @File  :SubForm10_GetUnitPrices.py
#! Python Version 3.7

from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
from tkinter import ttk
import xlrd
import xlwt
from xlutils.copy import copy
import os
import time
import datetime
import types
import openpyxl
from xlrd import xldate_as_tuple

'''
   Menu10
   功能模块：进行换算，由汇率资费表得到Unit Price
'''

monthList=['2018-10','2018-11','2018-12','2019-01','2019-02','2019-03','2019-04','2019-05','2019-06','2019-07','2019-08','2019-09','2019-10','2019-11','2019-12']
countryList=['AUD','CNY','HKD','IDR','INR','JPY','KRW','MYR','NZD','PHP','THB','TWD','VND','SGD']

def main():
    def selectExcelfile1():
        root.wm_attributes('-topmost',0)
        sfname = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        txt_ECBRateFile.insert(INSERT,sfname)
        root.wm_attributes('-topmost',1)


    def closeThisWindow():
        root.destroy()

    #第一个文件是ECB Rate File
    #第二个文件是APJ_FESTO_Services_Pricing_Mod.xls
    '''
       业务逻辑：
       A、Global Accounts Quarterly页
       1、step1 首先根据所选的年月，在第3行找到待汇总年月所在的列
       2、第1列Currency Name、第2列 ISO Code，这两列有一个为空则不处理，说明这一行不是数据
       3、找到1欧元=XXX美元，作为后面计算的基准
       4、汇率换算方式分两种：Currency Name里带base字样和不带base字样的
       5、只处理ASIA PACIFIC的即可，14个Country
       
    '''

    '''
       选择的ECB Rate File，是由虞琦给过来的文件，共两页
    '''
    def doProcess():
        
        #检查界面输入条件是否完整
        if checkForm()==False:
            tkinter.messagebox.showinfo('提示信息','请检查界面输入条件是否完备：\n1.Select Month \n2.Select ECB Rate File')
            return
        
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示信息','点确定按钮，开始处理......')
        root.wm_attributes('-topmost',1)
        sECBRateFile=txt_ECBRateFile.get()
        print(sECBRateFile)
        sECBRateFileName=sECBRateFile.split('/')[-1]

        dictList={}
        for i in range(len(monthList)-1):
            dictList[i]=monthList[i]
        sYM=dictList.get(lst_YM.current())
        wb1 = xlrd.open_workbook(filename=sECBRateFile)
        wb = openpyxl.load_workbook(sECBRateFile)
        ws=wb.worksheets[1]
        ws.title=sYM.replace('-','')                                                    #设置sheet name，把sYM里的yyyy-mm里的中短线去掉
        
        sheet1 = wb1.sheet_by_index(0)                                                  #Global Accounts Quarterly
        nrows1 = sheet1.nrows
        ncols1 = sheet1.ncols

        sheet2 = wb1.sheet_by_index(1)
        nrows2 = sheet2.nrows
        ncols2 = sheet2.ncols

        #step1 首先根据年月在Global Accounts Quarterly页在第3行找到换算年月所在的列
        iYMCol = 0
         
        for iCol in range(4,ncols1):
            sCellValue=sheet1.cell_value(2,iCol)
            sYMTemp=getYearMonth(sheet1.cell(2, iCol).ctype,sheet1.cell_value(2,iCol))
      
            
            if sYM==sYMTemp:
                iYMCol=iCol
                break
        if iYMCol==0:
            tkinter.messagebox.showinfo('提示信息','未在'+sECBRateFileName +'文件第1页第3行\n找到月份'+sYM +",请检查")
            return
        
            #step2 找到1欧元=XXX美金
        rateEuro = 0
        for i in range(1,nrows1):
            sCurrencyName = sheet1.cell_value(i,0)
            if sCurrencyName == "Euro - EUR base":
                rateEuro = sheet1.cell_value(i,iYMCol)
                break
        if rateEuro ==0 :
            tkinter.messagebox.showinfo('提示信息','未在'+sECBRateFileName +'文件第1页找到欧元所在的行,请检查，这里使用的查找信息：\nEuro - EUR base')
            return

        #step3 开始填写第2页的第2行
        ws.cell(row=2,column=2).value=sYM
        #根据ISO Code来判断
        sRate=""
        index=0
        for i in range(1,nrows1):
            sCurrencyName=sheet1.cell_value(i,0)
            sISOCode = sheet1.cell_value(i,1).strip()
            if sISOCode != "" and len(sISOCode)==3:
                if sISOCode in countryList:
                    index=sCurrencyName.find('base')
                    if index>0:
                        sRate = rateEuro/sheet1.cell_value(i,iYMCol)
                    else:
                        sRate=sheet1.cell_value(i,iYMCol)*rateEuro
                    for j in range(3,ncols2):
                        if sISOCode[0:2]==sheet2.cell_value(0,j):
                            ws.cell(row=2,column=j+1).value=sRate
                            break
        
        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time()))
        sPath=os.getcwd()+"\\Result\\"
        sReportName="7.ECBRate_"+sYM.replace('-','')+'_'+strTime+"_Style1.xlsx"
        wb.save(sPath+sReportName)

        print("Work is over.")
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示信息','处理完毕,生成文件：\n'+sPath+sReportName)
        root.wm_attributes('-topmost',1)

    

    '''
        ===================================函数=========================================
        名称：getYearMonth
        功能：xlrd从Excel读出的日期单元格的数据，不是以日期形式显示的，转化为yyyy-mm格式
        参数：ctype 单元格类型，3 为日期类型
              sCellValue 从单元格传过来的值
    '''
    def getYearMonth(ctype,sCellValue):
        sYM=""
        if ctype==3: 
            date = xlrd.xldate_as_tuple(sCellValue, 0)                   #转化为元组形式，如2019-02-01转化为：(2019, 2, 1, 0, 0, 0)
     
            sYear=str(date[0])
            sMonth=str(date[1])
            if len(sMonth)==1:
                sMonth='0'+sMonth
            sYM=sYear+'-'+sMonth
        return sYM
    '''
       =====================================函数========================================
       名称：checkForm
       功能：检查界面输入条件，在点Process按钮以前，是否已给出相应条件
    '''
    def checkForm():
        sYM = lst_YM.current()
        sECBRateFile = txt_ECBRateFile.get()

        if sECBRateFile == "":
            return False
        if  sYM == -1:
            return False
        
        return True

    #初始化
    root=Tk()

    #设置窗体标题
    root.title('Get Unit Pirce ---Step 7 ---Style 1')

    #设置窗口大小和位置
    root.geometry('660x300+430+220')

    #设置界面组件
    #组件第1行
    lbl_Month        = Label(root,text='Select Month:',width=16,justify = tkinter.RIGHT)
    sMonth           = tkinter.StringVar()
    lst_YM           = ttk.Combobox(root,width=20,textvariable=sMonth) #下拉列表框
    lst_YM['values'] = (monthList)


    #组件第2行
    lbl_SelectFile   = Label(root,text='Select ECB Rate File:',width=16,justify = tkinter.LEFT)
    txt_ECBRateFile  = Entry(root,bg='white',width=62)
    btn_browse1      = Button(root,text='Browse',width=8,command=selectExcelfile1)

    #组件第3行
    btn_process=Button(root,text='Process',width=10,command=doProcess)
    btn_exit=Button(root,text='Exit',width=10,command=closeThisWindow)

    lbl_Month.pack()
    lst_YM.pack()

    lbl_SelectFile.pack()
    txt_ECBRateFile.pack()
    btn_browse1.pack()

    btn_process.pack()
    btn_exit.pack() 

    lbl_SelectFile.place(x=50,y=60)
    txt_ECBRateFile.place(x=168,y=60)
    btn_browse1.place(x=550,y=56)


    lbl_Month.place(x=66,y=30)
    lst_YM.place(x=168,y=30)

    btn_process.place(x=250,y=130)
    btn_exit.place(x=360,y=130)

    root.mainloop() 

 
if __name__=="__main__":
    main()
