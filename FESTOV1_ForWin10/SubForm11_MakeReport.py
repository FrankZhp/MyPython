#! -*- coding utf-8 -*-
#! @Time  :2019/3/6 9:00
#! Author :Frank Zhang
#! @File  :SubForm9_MakeReport.py
#! Python Version 3.7
"""
#2个文件：
#1、已经计算了费用的单Sheet Excel,从第8步的文件获取
#2、待生成报表模板样例文件

#模块功能：按Price Table和Support分Sheet

【处理逻辑】：
 1、按Support分，Dispatch的分一类，非Dispatch的分一类
 2、按Price Table分，每月分一个Sheet
 3、再进一步处理，按月（Price Table）和Support Type分Sheet
 4、用xlrd读取上一工序生成的数据在一个sheet文件
"""

from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
from tkinter import ttk
import os
import time
import openpyxl
import xlrd


def main():
    def selectExcelFile():
        root.wm_attributes('-topmost',0)
        sfname = filedialog.askopenfilename(title='Please Select Excel File', filetypes=[('Excel', '*.xls'), ('All Files', '*')])
        #print(sfname)
        txt_ExcelFile.insert(INSERT,sfname)
        root.wm_attributes('-topmost',1)


    def closeThisWindow():
        root.destroy()
        

    def doProcess():
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示信息','点OK按钮，开始处理......')
        root.wm_attributes('-topmost',1)
        dictArea={0:'AU',1:'NZ',2:'CN',3:'HK',4:'TW',5:'SG',6:'TH',7:'ID',8:'MY',9:'VN',10:'PH'}
        dictPeriod={0:'2018Q11',1:'2018Q2',2:'2018Q3',3:'2018Q4',4:'2019Q1',5:'2019Q2',6:'2019Q3',7:'2019Q4'}

        sSourceFile=txt_ExcelFile.get()
        sPriceFile=os.getcwd()+"\\Source\\APJ_FESTO_Services_Pricing_FY19Q1.xls"
        sModFile=os.getcwd()+"\\Mod\\FESTO_Invoice_Summary_Style2_Mod.xlsx"

        wb1 = xlrd.open_workbook(filename=sSourceFile)
        wb2 = openpyxl.load_workbook(sModFile)

        sheet1 = wb1.sheet_by_index(1)
        nrows1 = sheet1.nrows                                               #Summary File max row's number
        ncols1 = sheet1.ncols                                               #Summary File max col's number

        sArea1=dictArea.get(lst_Area.current())

        sPeriod1=dictPeriod.get(lst_Period.current())
        
        # iRow2=1
        lst_SICode=[]
        dict_SICode={}
        lst_SICode2=[]

        #定义二维列表
        list_a=[]
        list_b=[]


        for iRow in range(1,nrows1):
            sSupportType=sheet1.cell_value(iRow,6)
            sPriceTable=sheet1.cell_value(iRow,56)
            list_b=[sSupportType,sPriceTable]
            if list_b not in list_a:
                list_a.append(list_b)


        while ['',''] in list_a:
            list_a.remove( ['',''])
        for list_c in list_a:

            sSupportType0=list_c[0]
            sPriceTable0=list_c[1]
            sNewSheetName=sSupportType0+sPriceTable0.replace("-","")

            ws=writeNewSheet(wb2,sNewSheetName)

            iRow2=1
            tmpRowNum=0
            lst_SICode=[]
            lst_SIItem=[]                    #记录SI Item No、SI Code、Unit Price
            lst_si=[]
            for iRow in range(1,nrows1):
                sTicketNo=sheet1.cell_value(iRow,0)
                sArea2=sheet1.cell_value(iRow,7).strip()

                sClosedDate=sheet1.cell_value(iRow,32).strip()
                sPeriod2=getPeriod(sClosedDate)
                sPriceTable=getPriceTable(sClosedDate)

                sSupportType2=sheet1.cell_value(iRow,6)
                sPriceTable2=sheet1.cell_value(iRow,56)

                if sSupportType0!=sSupportType2 or sPriceTable0!=sPriceTable2:
                    continue

                if sTicketNo!="" and sArea1==sArea2 and sPeriod1==sPeriod2:
                    iRow2=iRow2+1
                    for iCol in range(0,15):
                        ws.cell(row=iRow2,column=iCol+1,value=sheet1.cell_value(iRow,iCol))

                    for iCol in range(26,57):
                        ws.cell(row=iRow2,column=iCol+1,value=sheet1.cell_value(iRow,iCol))

                    sSIItemNo=sheet1.cell_value(iRow,52)
                    sSICode=sheet1.cell_value(iRow,53)
                    sPrice=sheet1.cell_value(iRow,54)
                    lst_SICode.append(sSICode)

                    #构建二维数组，存入SI Item No、SI Code、
                    lst_si=[sSIItemNo,sSICode,sPrice]
                    if lst_si not in lst_SIItem:
                        lst_SIItem.append(lst_si)

            dict = {}
            for key in lst_SICode:
                dict[key] = dict.get(key, 0) + 1

            iRow2=iRow2+3
            ws.cell(row=iRow2,column=1).value="Service Item No"
            ws.cell(row=iRow2,column=2).value="SI Code"
            ws.cell(row=iRow2,column=3).value="Unit Price"
            ws.cell(row=iRow2,column=4).value="Ticket"
            ws.cell(row=iRow2,column=5).value="TotalPrice"
            totalPrice=0
            subtotal =0
            for k,v in dict.items():
                sSICode=k
                iRow2=iRow2+1
                if FindSIItem(k,lst_SIItem,1)!=False:
                    tempRows=FindSIItem(k,lst_SIItem,1)
                        
                    sServiceItemNo=tempRows[0]
                    priceItem=tempRows[2]
                    ws.cell(row=iRow2,column=1).value=sServiceItemNo
                    ws.cell(row=iRow2,column=2).value=k
                    ws.cell(row=iRow2,column=3).value=priceItem
                    ws.cell(row=iRow2,column=4).value=v
                    if str(priceItem)!="":
                        subtotal=round(v*priceItem,2)                                         #int(v)*float(priceItem)
                    
                    totalPrice=totalPrice+subtotal
                    #print('Subtotal:'+str(subtotal)+'类型：')
                    ws.cell(row=iRow2,column=5).value=subtotal
                else:
                    print("Not Found:"+str(iRow2))

            iRow2=iRow2+1
            ws.cell(row=iRow2,column=1).value="Total"
            ws.cell(row=iRow2,column=5).value=round(totalPrice,2)

        #去掉mod页
        modsheeet=wb2["mod"]
        wb2.remove(modsheeet)

        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time()))
        sPath=os.getcwd()+"\\Result\\"
        sReportName="10.ComputePriceSyeStyle3_"+strTime+".xlsx"
        wb2.save(sPath+sReportName)

        print("Work is over.")
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示','处理完毕，生成文件：\n'+sPath+sReportName)
        root.wm_attributes('-topmost',1)

    def getPriceTable(sClosedDate):
        sPriceTable=""
        sYear=""
        sMonth=""
        if sClosedDate!="":
            list1=sClosedDate.split(" ")
            sDate=list1[0]
            list2=sDate.split("/")
            i=0
            for row in list2:
                if i==0:
                    sMonth=row
                #if i==1:
                #    sDay=row
                if i==2:
                    sYear=row
                i=i+1
            sPriceTable=sYear+"-"+sMonth     
        return sPriceTable
    def getPeriod(sClosedDate):
        sPeriod=""
        sYear=""
        sMonth=""
        sDay=""
        try:
            if sClosedDate!="":
                list1=sClosedDate.split(" ")
                sDate=list1[0]
                list2=sDate.split("/")
                i=0
                for row in list2:
                    if i==0:
                        sMonth=row
                    if i==1:
                        sDay=row
                    if i==2:
                        sYear=row
                    i=i+1
            dict={1:'Q1',2:'Q1',3:'Q1',4:'Q2',5:'Q2',6:'Q2',7:'Q3',8:'Q3',9:'Q3',10:'Q4',11:'Q4',12:'Q4'}
            sPeriod=sYear+dict.get(int(sMonth))
        except:
            print(sClosedDate)
        return sPeriod 

    def writeNewSheet(wb,sSheetName):
        ws5=wb.copy_worksheet(wb.worksheets[1])
        ws5.title=sSheetName
        return ws5
    
    def FindSIItem(target,array,searchCol):
        if not array:
            return False

        # 二维数组的行
        row = len(array)

        # 二维数组的列
        col = len(array[0])

        # 二层循环遍历二维数组
        for i in range(row):
            # 如果目标值等于数组中的值，则找到
            if target == array[i][searchCol]:
                return array[i]
        # 数组遍历结束后仍未找到
        return False
    
    #初始化
    root=Tk()

    #设置窗体标题
    root.title('Compute Price---Make Report-----Style 3-----Step10')

    #设置窗口大小和位置
    root.geometry('660x300+430+220')


    lbl_Summary=Label(root,text='Excel File:',width=20,justify = tkinter.RIGHT)
    txt_ExcelFile=Entry(root,bg='white',width=65)
    btn_browse1=Button(root,text='Browse',width=8,command=selectExcelFile)

    lbl_Area=Label(root,text="Select Country:",width=16,justify = tkinter.RIGHT)

    sArea = tkinter.StringVar()
    lst_Area=ttk.Combobox(root,width=20,textvariable=sArea) #下拉列表框
    lst_Area['values']=('AU','NZ','CN','HK','TW','SG','TH','ID','MY','VN','PH')
 

    lbl_Quarter=Label(root,text="Select Period:",justify = tkinter.RIGHT)

    #r = tkinter.StringVar()
    #r.set("Q1")
    
    #rdo_Quarter1 = tkinter.Radiobutton(root,
    #                                  variable = r,
    #                                  value = "Q1",
    #                                  text = "Q1")
    #rdo_Quarter2 = tkinter.Radiobutton(root,
    #                                  variable = r,
    #                                  value = "Q2",
    #                                  text = "Q2")
    #rdo_Quarter3 = tkinter.Radiobutton(root,
    #                                  variable = r,
    #                                  value = "Q3",
    #                                  text = "Q3")
    #rdo_Quarter4 = tkinter.Radiobutton(root,
    #                                  variable = r,
    #                                  value = "Q4",
    #                                  text = "Q4")
    sPeriod = tkinter.StringVar()
    lst_Period=ttk.Combobox(root,width=20,textvariable=sPeriod) #下拉列表框
    lst_Period['values']=('2018Q1','2018Q2','2018Q3','2018Q4','2019Q1','2019Q2','2019Q3','2019Q4')

    
    btn_process=Button(root,text='Process',width=8,command=doProcess)
    btn_exit=Button(root,text='Exit',width=8,command=closeThisWindow)
 

    lbl_Summary.pack()
    txt_ExcelFile.pack()
    btn_browse1.pack()


    lbl_Area.pack()

    lbl_Quarter.pack()
    #rdo_Quarter1.pack()
    #rdo_Quarter2.pack()
    #rdo_Quarter3.pack()
    #rdo_Quarter4.pack()
    lst_Period.pack()
    
    btn_process.pack()
    btn_exit.pack() 

    lbl_Summary.place(x=34,y=30)
    txt_ExcelFile.place(x=139,y=30)
    btn_browse1.place(x=550,y=26)

    lbl_Area.place(x=33,y=60)
    lst_Area.place(x=139,y=60)

    lbl_Quarter.place(x=300,y=60)
    #rdo_Quarter1.place(x=390,y=90)
    #rdo_Quarter2.place(x=430,y=90)
    #rdo_Quarter3.place(x=470,y=90)
    #rdo_Quarter4.place(x=510,y=90)
    lst_Period.place(x=385,y=60)
    
    btn_process.place(x=250,y=120)
    btn_exit.place(x=350,y=120)
 
 
    root.mainloop() 

 
if __name__=="__main__":
    main()
