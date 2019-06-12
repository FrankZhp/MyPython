from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
import xlrd
import xlwt
from xlutils.copy import copy
import os
import time
import openpyxl



def main():
    def selectExcelfile1():
        root.wm_attributes('-topmost',0)
        sfname = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        txt_ExcelFile.insert(INSERT,sfname)
        root.wm_attributes('-topmost',1)

    def selectExcelfile2():
        root.wm_attributes('-topmost',0)
        sfname = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '*.xls'), ('All Files', '*')])
        txt_ClosedTicketFile.insert(INSERT,sfname)
        root.wm_attributes('-topmost',1)

    def closeThisWindow():
        root.destroy()

    def doProcess():
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示信息','点OK按钮，开始处理......')
        root.wm_attributes('-topmost',1)
        sExcelFile=txt_ExcelFile.get()
        sClosedFile=txt_ClosedTicketFile.get()
        #wb1 = xlrd.open_workbook(filename=sMergedFile, formatting_info=True)
        wb1 = openpyxl.load_workbook(sExcelFile)
        wb2 = xlrd.open_workbook(filename=sClosedFile)

        ws = wb1.worksheets[1]
        #sheet1 = wb1.sheet_by_index(1)   
        #nrows1 = sheet1.nrows
        nrows1  = ws.max_row

        sheet2 = wb2.sheet_by_index(0)   
        nrows2 = sheet2.nrows
        ncols2 = sheet2.ncols
        cols2 = sheet2.col_values(1)                #第2列

        
        for i in range(1,nrows1+1):
            #sCellValue=sheet1.cell_value(i,0)
            sCellValue = ws.cell(i,1).value
            sTicketNo = sCellValue.strip()
            if sTicketNo in cols2:
                j=cols2.index(sTicketNo)
                #ws.write(i,19,sheet2.cell_value(j,9))                         #补充Main Description
                ws.cell(row=i,column=20).value=sheet2.cell_value(j,9)
                
            else:
                print('Not Found:'+sTicketNo)
       
        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) 
        sPath=os.getcwd()+"\\Result\\"
        sReportName="11.Report_"+strTime+".xlsx"
        wb1.save(sPath+sReportName)        
       

        print("Work is over.")
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示信息','处理完毕，生成文件：\n'+sPath+sReportName)
        root.wm_attributes('-topmost',1)
    
    #初始化
    root=Tk()

    #设置窗体标题
    root.title('Mapping---Step 11')

    #设置窗口大小和位置
    root.geometry('660x300+430+220')


    label1=Label(root,text='Select Excel File:')
    txt_ExcelFile=Entry(root,bg='white',width=67)
    btn_browse1=Button(root,text='Browse',width=8,command=selectExcelfile1)

    label2=Label(root,text='Closed Tickets File:')
    txt_ClosedTicketFile=Entry(root,bg='white',width=67)
    btn_browse2=Button(root,text='Browse',width=8,command=selectExcelfile2)
    
    btn_process=Button(root,text='Process',width=8,command=doProcess)
    btn_exit=Button(root,text='Exit',width=8,command=closeThisWindow)
 

    label1.pack()
    txt_ExcelFile.pack()
    btn_browse1.pack()

    label2.pack()
    txt_ClosedTicketFile.pack()
    btn_browse2.pack()
    
    btn_process.pack()
    btn_exit.pack() 

    label1.place(x=30,y=30)
    txt_ExcelFile.place(x=138,y=30)
    btn_browse1.place(x=550,y=26)

    label2.place(x=30,y=60)
    txt_ClosedTicketFile.place(x=138,y=60)
    btn_browse2.place(x=550,y=56)
    
    btn_process.place(x=250,y=100)
    btn_exit.place(x=350,y=100)
 
 
    root.mainloop() 

 
if __name__=="__main__":
    main()
