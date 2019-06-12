#! -*- coding utf-8 -*-
#! @Time  :2019/3/1 9:44
#! Author :Frank Zhang
#! @File  :Python_Blank.py
#! Python Version 3.7

#模块功能：读取某个文件夹下的所有FESTO Excel文件，读取每个文件里的所有Sheet，按照模板文件的样例，读取相应列，把数据合并到一个Excel文件中
#业务逻辑说明
#1.把模板文件的第一列读入list中,读取的FESTO文件，个别的可能与模板文件的列以及顺序不一样，那么按照模板文件的列及顺序，合并到一个新的文件中。
#2.模板文件里没有的列则不读取


from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
from tkinter.filedialog import askdirectory
import xlrd
import xlwt
import os
import openpyxl
import datetime
import time

 

def main():
    #Get Directory
    def getDirectory():
        path_ = askdirectory()
        txt_Directory.insert(INSERT,path_)

    def selectExcelfile():
        sfname = filedialog.askopenfilename(title='Select Excel File', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        txt_ExcelModFile.insert(INSERT,sfname)
        
    def getModColumn():
        #list，把列名放在其中
        modlist=['Ticket No','Period','Invoice	FESTO PO','Invoice Amount','INV_Currency','Remarks']		  		

        
    def closeThisWindow():
        root.destroy()

    def doProcess():
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示','开始处理______')
        root.wm_attributes('-topmost',1)
        sPath=txt_Directory.get()
        #print(sPath)
        
        filenames = os.listdir(sPath)
        #print(filenames)
        
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(u'Sheet1')
        #为去重，定义一个list，遍历时，判断遍历到的TicketNo是否在此list中
        #如果不在，则增加到这个list中，并填写到结果文件中
        last_List=[]                            

        #iRow记录写的Excel的行号，第一行是从0开始，所以，初始化为-1
        iRow2=-1
        iCol2=-1

        #读取Excel模板文件，把模板文件里的的第一行的列名作为结果文件的列名
        modlist=['Ticket No','Period','Invoice','FESTO PO','Invoice Amount','INV_Currency','Remarks']	

        #填写结果文件的第一行
        iRow2=0
        for i in range(len(modlist)):
            worksheet.write(iRow2,i,modlist[i])

        for iFile, filename in enumerate(filenames):
            sFileName=filename
                
            #wb = openpyxl.load_workbook(sPath+'\\'+sFileName)
            wb = xlrd.open_workbook(filename=sPath+'\\'+sFileName)
            
            sheet = wb.sheet_by_index(2)
            rows = sheet.row_values(1)   # 获取第一行内容
            cols = sheet.col_values(17)  #获取第17列(R列)的内容
            max_row=len(cols)
            max_col=len(rows)
            #判断这个Sheet的第一行第一个单元格是否是Ticket No.，如果不是，则不读取这一页
            sColumn=sheet.cell(1,17).value
            #print("第一个单元格："+sColumn1)
            if sColumn=="Ticket No":
                pass
            else:
                continue
                
      
            #第一列关键字，如果重复则去掉
            old_List=sheet.col_values(17)
                
            #第一行是列名，从第二行开始
            for iRow1 in range(2,max_row):
                 
                for iCol1 in range(6):
                    if iCol1==0:
                        try:
                            sTicketNo=sheet.cell(iRow1,iCol1+17).value
                            if sTicketNo.strip()=="":
                                pass
                            else:
                                
                                if old_List[iRow1] in last_List:                     #如果已有，则退出for循环，不增加重复数据
                                    break                                   
                                else:
                                    iRow2=iRow2+1
                                    last_List.append(old_List[iRow1])                   #把没有关键字增加到列表中
                                    worksheet.write(iRow2,iCol1,sheet.cell(iRow1,iCol1+17).value)
                        except:
                            print("error1")
                       
                    else:
                        try:
                            if sTicketNo.strip()=="":
                                pass
                            else:
                                worksheet.write(iRow2,iCol1,sheet.cell(iRow1,iCol1+17).value)
                        except:
                            print('error2')

        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) 
        sPath=os.getcwd()+"\\Result\\"
        sReportName="5.MergeInvoice_"+strTime+".xls"
        workbook.save(sPath+sReportName)

        print("Work is over.")
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示','处理完毕,生成文件：\n'+sPath+sReportName)
        root.wm_attributes('-topmost',1)

    #初始化
    root=Tk()

    #设置窗体标题
    root.title('Merge Invoice files---Step 5')

    #设置窗口大小和位置
    root.geometry('660x300+430+220')

    label1=Label(root,text='Select Path:',justify = tkinter.RIGHT)
    txt_Directory=Entry(root,bg='white',width=70)
    btn_Browse1=Button(root,text='Browse',width=8,command=getDirectory)
    
    btn_DoProcess=Button(root,text='Process',width=8,command=doProcess)
    btn_Exit=Button(root,text='Exit',width=8,command=closeThisWindow)
 

    label1.pack()
    txt_Directory.pack()

    btn_Browse1.pack()
 
    
    btn_DoProcess.pack()
    btn_Exit.pack() 

    label1.place(x=30,y=30)
    txt_Directory.place(x=108,y=30)
    btn_Browse1.place(x=550,y=26)

    btn_DoProcess.place(x=200,y=100)
    btn_Exit.place(x=300,y=100)
 
    root.mainloop() 

def getFileNames(path):
    filenames = os.listdir(path)
    for i, filename in enumerate(filenames):
         if i==0:
            iSpecialFile=i+1
            sFileName=filename

def getSheetNames(path,sFileName):
    wb = openpyxl.load_workbook(path+'\\'+sFileName)
    # 获取workbook中所有的表格
    sheets = wb.sheetnames

    # 循环遍历所有sheet
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        
 
if __name__=="__main__":
    main()
