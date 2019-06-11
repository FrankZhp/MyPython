# -*- coding: utf-8 -*-

"""
Module implementing MergeInvoice.
模块功能：读取某个文件夹下的所有FESTO Excel文件，读取每个文件里的所有Sheet，按照模板文件的样例，读取相应列，把数据合并到一个Excel文件中
业务逻辑说明
1.把模板文件的第一列读入list中,读取的FESTO文件，个别的可能与模板文件的列以及顺序不一样，那么按照模板文件的列及顺序，合并到一个新的文件中。
2.模板文件里没有的列则不读取

"""

from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog
from PyQt5 import  QtWidgets

from Ui_SubForm5_MergeInvoice import Ui_Form

import xlrd
import xlwt
import os
import openpyxl
import datetime
import time

class MergeInvoice(QWidget, Ui_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(MergeInvoice, self).__init__(parent)
        self.setupUi(self)
    
    @pyqtSlot()
    def on_btn_Process_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        reply = QMessageBox.information(self,                         #使用infomation信息框
                                    "提示信息",
                                    "点Yes按钮,开始处理......",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply ==QMessageBox.No:
            return
        
        sPath=self.txt_Directory.get()
        #print(sPath)
        
        filenames = os.listdir(sPath)
        #print(filenames)
        
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(u'Sheet1')
        #为去重，定义一个list，遍历时，判断遍历到的TicketNo是否在此list中
        #如果不在，则增加到这个list中，并填写到结果文件中
        last_List=[]                            

        #iRow记录写的Excel的行号，第一行是从0开始，所以，初始化为-1
        iRow2=-1
        iCol2=-1

        #读取Excel模板文件，把模板文件里的的第一行的列名作为结果文件的列名
        modlist=['Ticket No','Period','Invoice','FESTO PO','Invoice Amount','INV_Currency','Remarks']	

        #填写结果文件的第一行
        iRow2=0
        for i in range(len(modlist)):
            worksheet.write(iRow2,i,modlist[i])

        for iFile, filename in enumerate(filenames):
            sFileName=filename
                
            #wb = openpyxl.load_workbook(sPath+'\\'+sFileName)
            wb = xlrd.open_workbook(filename=sPath+'\\'+sFileName)
            
            sheet = wb.sheet_by_index(2)
            rows = sheet.row_values(1)   # 获取第一行内容
            cols = sheet.col_values(17)  #获取第17列(R列)的内容
            max_row=len(cols)
            max_col=len(rows)
            #判断这个Sheet的第一行第一个单元格是否是Ticket No.，如果不是，则不读取这一页
            sColumn=sheet.cell(1,17).value
            #print("第一个单元格："+sColumn1)
            if sColumn=="Ticket No":
                pass
            else:
                continue
                
      
            #第一列关键字，如果重复则去掉
            old_List=sheet.col_values(17)
                
            #第一行是列名，从第二行开始
            for iRow1 in range(2,max_row):
                 
                for iCol1 in range(6):
                    if iCol1==0:
                        try:
                            sTicketNo=sheet.cell(iRow1,iCol1+17).value
                            if sTicketNo.strip()=="":
                                pass
                            else:
                                
                                if old_List[iRow1] in last_List:                     #如果已有，则退出for循环，不增加重复数据
                                    break                                   
                                else:
                                    iRow2=iRow2+1
                                    last_List.append(old_List[iRow1])                   #把没有关键字增加到列表中
                                    worksheet.write(iRow2,iCol1,sheet.cell(iRow1,iCol1+17).value)
                        except:
                            print("error1")
                       
                    else:
                        try:
                            if sTicketNo.strip()=="":
                                pass
                            else:
                                worksheet.write(iRow2,iCol1,sheet.cell(iRow1,iCol1+17).value)
                        except:
                            print('error2')

        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) 
        sPath=os.getcwd()+"\\Result\\"
        sReportName="5.MergeInvoice_"+strTime+".xls"
        workbook.save(sPath+sReportName)
        
        print("Work is over.")
        QMessageBox.information(self,"提示信息","处理完毕，存放位置：当前目录\\Result,\n文件名称:"+sReportName,QMessageBox.Yes)

    
    @pyqtSlot()
    def on_btn_Exit_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        reply = QMessageBox.question(self, "标题", "确定要退出吗?", 
                   QMessageBox.Yes|QMessageBox.No)
         
        if reply == QMessageBox.Yes:
            self.close()
    
    @pyqtSlot()
    def on_btn_Browse_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        directory1  = QFileDialog.getExistingDirectory(self, "选择文件", "./")
        self.txt_File.setPlainText(directory1 )
    
    def InitForm(self):
        self.txt_Directory.setPlainText("")   
        
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    subForm5 = MergeInvoice()
    subForm5.show()
    sys.exit(app.exec_())
