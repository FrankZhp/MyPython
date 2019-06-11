# -*- coding: utf-8 -*-

"""
Module implementing GetUnitPricesStyle1.
功能模块：进行换算，由汇率资费表得到Unit Price

    #第一个文件是ECB Rate File
    #第二个文件是APJ_FESTO_Services_Pricing_Mod.xls
    '''
       业务逻辑：
       A、Global Accounts Quarterly页
       1、step1 首先根据所选的年月，在第3行找到待汇总年月所在的列
       2、第1列Currency Name、第2列 ISO Code，这两列有一个为空则不处理，说明这一行不是数据
       3、找到1欧元=XXX美元，作为后面计算的基准
       4、汇率换算方式分两种：Currency Name里带base字样和不带base字样的
       5、只处理ASIA PACIFIC的即可，14个Country
       
    '''

    '''
       选择的ECB Rate File，是由虞琦给过来的文件，共两页
    '''
"""

from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog
from PyQt5 import  QtWidgets

from Ui_SubForm7_GetUnitPrices import Ui_Form

import xlrd
import xlwt
from xlutils.copy import copy
import os
import time
import datetime
import types
import openpyxl
from xlrd import xldate_as_tuple

monthList=[]
countryList=['AUD','CNY','HKD','IDR','INR','JPY','KRW','MYR','NZD','PHP','THB','TWD','VND','SGD']

class GetUnitPricesStyle1(QWidget, Ui_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(GetUnitPricesStyle1, self).__init__(parent)
        self.setupUi(self)
        
        monthList = self.InitMonth()
        self.cbo_Month.addItem('')
        self.cbo_Month.addItems(monthList)
    
    @pyqtSlot()
    def on_btn_Process_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        reply = QMessageBox.information(self,                         #使用infomation信息框
                                    "提示信息",
                                    "点Yes按钮,开始处理......",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply ==QMessageBox.No:
            return
        sECBRateFile=self.txt_ECBRateFile.toPlainText()
         
        sECBRateFileName=sECBRateFile.split('/')[-1]
         
        #dictList={}
        #for i in range(len(monthList)-1):
        #    dictList[i]=monthList[i]
        #sYM=dictList.get(self.cbo_Month.currentText())
        sYM=self.cbo_Month.currentText()
        wb1 = xlrd.open_workbook(filename=sECBRateFile)
        wb = openpyxl.load_workbook(sECBRateFile)
        ws=wb.worksheets[1]
        
        ws.title=sYM.replace('-','')                                                    #设置sheet name，把sYM里的yyyy-mm里的中短线去掉
        
        sheet1 = wb1.sheet_by_index(0)                                                  #Global Accounts Quarterly
        nrows1 = sheet1.nrows
        ncols1 = sheet1.ncols

        sheet2 = wb1.sheet_by_index(1)
        nrows2 = sheet2.nrows
        ncols2 = sheet2.ncols

        #step1 首先根据年月在Global Accounts Quarterly页在第3行找到换算年月所在的列
        iYMCol = 0
         
        for iCol in range(4,ncols1):
            sCellValue=sheet1.cell_value(2,iCol)
            sYMTemp=self.getYearMonth(sheet1.cell(2, iCol).ctype,sheet1.cell_value(2,iCol))
      
            
            if sYM==sYMTemp:
                iYMCol=iCol
                break
        if iYMCol==0:
            QMessageBox.information(self,"提示信息",'未在'+sECBRateFileName +'文件第1页第3行\n找到月份'+sYM +",请检查",QMessageBox.Yes)
            
            return
        
            #step2 找到1欧元=XXX美金
        rateEuro = 0
        for i in range(1,nrows1):
            sCurrencyName = sheet1.cell_value(i,0)
            if sCurrencyName == "Euro - EUR base":
                rateEuro = sheet1.cell_value(i,iYMCol)
                break
        if rateEuro ==0 :
            QMessageBox.information(self,"提示信息",'未在'+sECBRateFileName +'文件第1页找到欧元所在的行,请检查，这里使用的查找信息：\nEuro - EUR base',QMessageBox.Yes)

            return

        #step3 开始填写第2页的第2行
        ws.cell(row=2,column=2).value=sYM
        #根据ISO Code来判断
        sRate=""
        index=0
        for i in range(1,nrows1):
            sCurrencyName=sheet1.cell_value(i,0)
            sISOCode = sheet1.cell_value(i,1).strip()
            if sISOCode != "" and len(sISOCode)==3:
                if sISOCode in countryList:
                    index=sCurrencyName.find('base')
                    if index>0:
                        sRate = rateEuro/sheet1.cell_value(i,iYMCol)
                    else:
                        sRate=sheet1.cell_value(i,iYMCol)*rateEuro
                    for j in range(3,ncols2):
                        if sISOCode[0:2]==sheet2.cell_value(0,j):
                            ws.cell(row=2,column=j+1).value=sRate
                            break
        
        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time()))
        sPath=os.getcwd()+"\\Result\\"
        sReportName="7.ECBRate_"+sYM.replace('-','')+'_'+strTime+"_Style1.xlsx"
        wb.save(sPath+sReportName)    
        
        print("Work is over.")
        QMessageBox.information(self,"提示信息","处理完毕，存放位置：当前目录\\Result,\n文件名称:"+sReportName,QMessageBox.Yes)
    
    @pyqtSlot()
    def on_btn_Exit_clicked(self):
        """
        Slot documentation goes here.
        """
        reply = QMessageBox.question(self, "标题", "确定要退出吗?", 
                   QMessageBox.Yes|QMessageBox.No)
         
        if reply == QMessageBox.Yes:
            self.close()
    
    @pyqtSlot()
    def on_btn_Browse_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        #raise NotImplementedError
        fileName1, filetype = QFileDialog.getOpenFileName(self, "选择文件", "./", "Excel Files (*.xlsx)")
        print(fileName1,filetype)
        #self.txt_File.setText=fileName1
        self.txt_ECBRateFile.setPlainText(fileName1)
    
    def InitForm(self):
        self.txt_ECBRateFile.setPlainText("")
        
        
    '''
        ===================================函数=========================================
        名称：getYearMonth
        功能：xlrd从Excel读出的日期单元格的数据，不是以日期形式显示的，转化为yyyy-mm格式
        参数：ctype 单元格类型，3 为日期类型
              sCellValue 从单元格传过来的值
    '''
    def getYearMonth(self, ctype,sCellValue):
        sYM=""
        if ctype==3: 
            date = xlrd.xldate_as_tuple(sCellValue, 0)                   #转化为元组形式，如2019-02-01转化为：(2019, 2, 1, 0, 0, 0)
     
            sYear=str(date[0])
            sMonth=str(date[1])
            if len(sMonth)==1:
                sMonth='0'+sMonth
            sYM=sYear+'-'+sMonth
        return sYM
    '''
       =====================================函数========================================
       名称：checkForm
       功能：检查界面输入条件，在点Process按钮以前，是否已给出相应条件
    '''
    def checkForm(self):
        sYM = self.cbo_Month.currentText()
        sECBRateFile = self.txt_ECBRateFile.toPlainText()

        if sECBRateFile == "":
            return False
        if  sYM == -1:
            return False
        
        return True
    """
       名称:InitMonth
        功能：初始化MonthList
        配置文件：./Doc/config.xls
    """
    def InitMonth(self):
        sPath = os.getcwd()+"\\Config\\"
        sConfigFile = "config.xls"
        wb1 = xlrd.open_workbook(filename=sPath+sConfigFile)    
        sheet1 = wb1.sheet_by_index(0)
        col1=sheet1.col_values(1)
        del col1[0]                               #去掉第1行的列名
        return col1
        
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    subForm7 = GetUnitPricesStyle1()
    subForm7.show()
    sys.exit(app.exec_())
